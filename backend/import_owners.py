from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import load_workbook
from sqlmodel import Session, select

from app.database import engine
from app.models import DuesPaymentMethod, Owner


def normalize_header(value: object) -> str:
    if value is None:
        return ""
    return "".join(ch for ch in str(value).strip().lower() if ch.isalnum())


def normalize_value(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text if text else None


def map_headers(header_row: tuple[object, ...]) -> dict[str, int]:
    aliases = {
        "unit": "unit",
        "unitnumber": "unit",
        "unit#": "unit",
        "name": "owner_one_full_name",
        "ownername": "owner_one_full_name",
        "ownersname": "owner_one_full_name",
        "fullname": "owner_one_full_name",
        "owner1name": "owner_one_full_name",
        "owner1fullname": "owner_one_full_name",
        "ownerone": "owner_one_full_name",
        "owneronefullname": "owner_one_full_name",
        "owneronefull_name": "owner_one_full_name",
        "owneronefull": "owner_one_full_name",
        "owneroneemail": "owner_one_email",
        "owner1email": "owner_one_email",
        "owner1emailaddress": "owner_one_email",
        "owneroneemailaddress": "owner_one_email",
        "email": "owner_one_email",
        "emailaddress": "owner_one_email",
        "owneronephone": "owner_one_phone",
        "owner1phone": "owner_one_phone",
        "owner1phonenumber": "owner_one_phone",
        "owneronephonenumber": "owner_one_phone",
        "phone": "owner_one_phone",
        "phonenumber": "owner_one_phone",
        "owneronemailingaddress": "owner_one_mailing_address",
        "owner1mailingaddress": "owner_one_mailing_address",
        "owner1mailing": "owner_one_mailing_address",
        "owneronemailing": "owner_one_mailing_address",
        "mailingaddress": "owner_one_mailing_address",
        "address": "owner_one_mailing_address",
        "owner2name": "owner_two_full_name",
        "owner2fullname": "owner_two_full_name",
        "ownertwoname": "owner_two_full_name",
        "ownertwofullname": "owner_two_full_name",
        "owner2email": "owner_two_email",
        "ownertwoemail": "owner_two_email",
        "owner2emailaddress": "owner_two_email",
        "ownertwoemailaddress": "owner_two_email",
        "owner2phone": "owner_two_phone",
        "ownertwophone": "owner_two_phone",
        "owner2phonenumber": "owner_two_phone",
        "ownertwophonenumber": "owner_two_phone",
        "owner2mailingaddress": "owner_two_mailing_address",
        "ownertwomailingaddress": "owner_two_mailing_address",
        "owner2mailing": "owner_two_mailing_address",
        "ownertwomailing": "owner_two_mailing_address",
    }
    mapped: dict[str, int] = {}
    for index, cell in enumerate(header_row):
        canonical = aliases.get(normalize_header(cell))
        if canonical is not None and canonical not in mapped:
            mapped[canonical] = index
    missing = [
        required
        for required in ("unit", "owner_one_full_name")
        if required not in mapped
    ]
    if missing:
        raise ValueError(
            f"Missing required columns: {', '.join(missing)}. "
            "Expected headers like unit and owner 1 name."
        )
    return mapped


def find_header_row(rows: list[tuple[object, ...]]) -> tuple[int, dict[str, int]]:
    max_scan = min(len(rows), 20)
    for index in range(max_scan):
        try:
            header_map = map_headers(rows[index])
            return index, header_map
        except ValueError:
            continue
    raise ValueError(
        "Could not detect header row with required columns. "
        "Expected columns like unit and owner 1 name."
    )


def get_row_value(
    row: tuple[object, ...], header_map: dict[str, int], key: str
) -> str | None:
    index = header_map.get(key)
    if index is None or index >= len(row):
        return None
    return normalize_value(row[index])


def import_owners(
    file_path: Path, sheet_name: str | None, replace_all: bool = False
) -> tuple[int, int, int, int]:
    workbook = load_workbook(filename=file_path, data_only=True)
    sheet = workbook[sheet_name] if sheet_name else workbook.active

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Spreadsheet is empty")

    header_row_index, header_map = find_header_row(rows)
    created = 0
    updated = 0
    deleted = 0
    skipped = 0
    imported_units: set[str] = set()

    with Session(engine) as session:
        for row in rows[header_row_index + 1 :]:
            unit = get_row_value(row, header_map, "unit")
            owner_one_full_name = get_row_value(row, header_map, "owner_one_full_name")
            owner_one_email = get_row_value(row, header_map, "owner_one_email")
            owner_one_phone = get_row_value(row, header_map, "owner_one_phone")
            owner_one_mailing_address = get_row_value(
                row, header_map, "owner_one_mailing_address"
            )

            owner_two_full_name = get_row_value(row, header_map, "owner_two_full_name")
            owner_two_email = get_row_value(row, header_map, "owner_two_email")
            owner_two_phone = get_row_value(row, header_map, "owner_two_phone")
            owner_two_mailing_address = get_row_value(
                row, header_map, "owner_two_mailing_address"
            )

            if unit is None or owner_one_full_name is None:
                skipped += 1
                continue

            imported_units.add(unit)

            existing = session.exec(
                select(Owner).where(Owner.unit_number == unit)
            ).first()
            if existing is None:
                owner = Owner(
                    unit_number=unit,
                    owner_one_full_name=owner_one_full_name,
                    owner_one_email=owner_one_email,
                    owner_one_phone=owner_one_phone,
                    owner_one_mailing_address=owner_one_mailing_address,
                    owner_two_full_name=owner_two_full_name,
                    owner_two_email=owner_two_email,
                    owner_two_phone=owner_two_phone,
                    owner_two_mailing_address=owner_two_mailing_address,
                    dues_payment_method=DuesPaymentMethod.CHECK,
                    active=True,
                )
                session.add(owner)
                created += 1
            else:
                existing.owner_one_full_name = owner_one_full_name
                existing.owner_one_email = owner_one_email
                existing.owner_one_phone = owner_one_phone
                existing.owner_one_mailing_address = owner_one_mailing_address
                existing.owner_two_full_name = owner_two_full_name
                existing.owner_two_email = owner_two_email
                existing.owner_two_phone = owner_two_phone
                existing.owner_two_mailing_address = owner_two_mailing_address
                session.add(existing)
                updated += 1

        if replace_all:
            existing_units = set(session.exec(select(Owner.unit_number)).all())
            units_to_delete = existing_units.difference(imported_units)
            for unit_to_delete in units_to_delete:
                owner = session.exec(
                    select(Owner).where(Owner.unit_number == unit_to_delete)
                ).first()
                if owner is not None:
                    session.delete(owner)
                    deleted += 1

        session.commit()

    return created, updated, deleted, skipped


def main() -> None:
    parser = argparse.ArgumentParser(description="Import owner data from an Excel file")
    parser.add_argument("excel_file", help="Absolute or relative path to .xlsx file")
    parser.add_argument("--sheet", default=None, help="Optional sheet name")
    parser.add_argument(
        "--replace-all",
        action="store_true",
        help="Delete owners not present in the spreadsheet after import",
    )
    args = parser.parse_args()

    file_path = Path(args.excel_file).expanduser().resolve()
    if not file_path.exists():
        raise FileNotFoundError(f"File not found: {file_path}")

    created, updated, deleted, skipped = import_owners(
        file_path=file_path, sheet_name=args.sheet, replace_all=args.replace_all
    )
    print(
        "Owner import complete "
        f"(created={created}, updated={updated}, deleted={deleted}, skipped={skipped})"
    )


if __name__ == "__main__":
    main()
