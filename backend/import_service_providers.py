from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import load_workbook
from sqlmodel import Session, select

from app.database import engine
from app.models import ServiceProvider


def normalize_header(value: object) -> str:
    if value is None:
        return ""
    return "".join(ch for ch in str(value).strip().lower() if ch.isalnum())


def normalize_value(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text if text else None


def map_headers(header_row: tuple[object, ...]) -> dict[str, int]:
    aliases = {
        "company": "company_name",
        "companyname": "company_name",
        "providername": "company_name",
        "vendor": "company_name",
        "vendorname": "company_name",
        "contact": "contact_name",
        "contactname": "contact_name",
        "phone": "phone",
        "phonenumber": "phone",
        "email": "email",
        "emailaddress": "email",
        "category": "service_category",
        "service": "service_category",
        "servicecategory": "service_category",
        "notes": "notes",
    }
    mapped: dict[str, int] = {}
    for index, cell in enumerate(header_row):
        canonical = aliases.get(normalize_header(cell))
        if canonical is not None and canonical not in mapped:
            mapped[canonical] = index

    if "company_name" not in mapped:
        raise ValueError(
            "Missing required column: company name. "
            "Expected headers like company, vendor, or provider name."
        )
    return mapped


def find_header_row(rows: list[tuple[object, ...]]) -> tuple[int, dict[str, int]]:
    max_scan = min(len(rows), 20)
    for index in range(max_scan):
        try:
            header_map = map_headers(rows[index])
            return index, header_map
        except ValueError:
            continue
    raise ValueError(
        "Could not detect header row with company name column in first 20 rows."
    )


def import_service_providers(
    file_path: Path, sheet_name: str | None
) -> tuple[int, int, int]:
    workbook = load_workbook(filename=file_path, data_only=True)
    sheet = workbook[sheet_name] if sheet_name else workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Spreadsheet is empty")

    header_row_index, header_map = find_header_row(rows)
    created = 0
    updated = 0
    skipped = 0

    with Session(engine) as session:
        for row in rows[header_row_index + 1 :]:
            company_name = normalize_value(row[header_map["company_name"]])
            contact_name = (
                normalize_value(row[header_map["contact_name"]])
                if "contact_name" in header_map
                else None
            )
            phone = (
                normalize_value(row[header_map["phone"]])
                if "phone" in header_map
                else None
            )
            email = (
                normalize_value(row[header_map["email"]])
                if "email" in header_map
                else None
            )
            service_category = (
                normalize_value(row[header_map["service_category"]])
                if "service_category" in header_map
                else None
            )
            notes = (
                normalize_value(row[header_map["notes"]])
                if "notes" in header_map
                else None
            )

            if company_name is None:
                skipped += 1
                continue

            existing = session.exec(
                select(ServiceProvider).where(
                    ServiceProvider.company_name == company_name
                )
            ).first()
            if existing is None:
                provider = ServiceProvider(
                    company_name=company_name,
                    contact_name=contact_name,
                    email=email,
                    phone=phone,
                    service_category=service_category,
                    notes=notes,
                    active=True,
                )
                session.add(provider)
                created += 1
            else:
                existing.contact_name = contact_name
                existing.email = email
                existing.phone = phone
                existing.service_category = service_category
                existing.notes = notes
                existing.active = True
                session.add(existing)
                updated += 1

        session.commit()

    return created, updated, skipped


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Import service provider data from an Excel file"
    )
    parser.add_argument("excel_file", help="Absolute or relative path to .xlsx file")
    parser.add_argument("--sheet", default=None, help="Optional sheet name")
    args = parser.parse_args()

    file_path = Path(args.excel_file).expanduser().resolve()
    if not file_path.exists():
        raise FileNotFoundError(f"File not found: {file_path}")

    created, updated, skipped = import_service_providers(
        file_path=file_path, sheet_name=args.sheet
    )
    print(
        "Service provider import complete "
        f"(created={created}, updated={updated}, skipped={skipped})"
    )


if __name__ == "__main__":
    main()
